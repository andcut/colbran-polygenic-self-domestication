#!/usr/bin/env python3
"""Run the cognitive comparison for IQ and EA4.

This script uses the same downstream choices as the GFP and schizophrenia
runner:

- intersect to AADR 1240k variants
- keep genome-wide-significant loci at p < 1e-8
- clump on modern-European AADR samples with PLINK at 200 kb and r^2 > 0.4
- run Colbran-style sign permutations across AFR, EAS, SAS, SAM, and NonEur

The trait side differs slightly across the two inputs:

- IQ uses full summary statistics from Savage et al. 2018
- EA4 uses the lead-hit supplementary table from Lee et al. 2018
"""

from __future__ import annotations

import argparse
import csv
import io
import math
import zipfile
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

from run_extension import (
    AADR_GENO,
    AADR_IND,
    AADR_SNP,
    POLYGENIC_SCRIPT,
    SAMPLE_INFO_DIR,
    Hit,
    aggregate_variant_counts,
    build_matched_hits,
    build_plink_reference,
    clean_value,
    deduplicate_hits,
    float_value,
    load_aadr_ind,
    load_aadr_variant_ids,
    load_eur_reference_indices,
    load_target_variants,
    normalize_allele,
    parse_region_groups,
    run_joint_sensitivity,
    run_plink_clump,
    run_sign_suite,
    write_admixture_files,
    write_counts_files,
    write_hits,
    write_match_summary,
)


ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "data" / "raw"
RESULTS_DIR = ROOT / "results" / "cognitive_comparison_100k"
IQ_SUMSTATS = RAW_DIR / "full_sumstats" / "SavageJansen_IntMeta_sumstats.zip"
EA4_XLSX = RAW_DIR / "gwas" / "ea4_supp_tables.xlsx"
TRAITS = ["iq", "ea4"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Run IQ and EA4 through the same Colbran-style pipeline used "
            "for GFP and schizophrenia."
        )
    )
    parser.add_argument("--output-dir", type=Path, default=RESULTS_DIR)
    parser.add_argument("--sample-info-dir", type=Path, default=SAMPLE_INFO_DIR)
    parser.add_argument("--aadr-ind", type=Path, default=AADR_IND)
    parser.add_argument("--aadr-snp", type=Path, default=AADR_SNP)
    parser.add_argument("--aadr-geno", type=Path, default=AADR_GENO)
    parser.add_argument("--iq-sumstats", type=Path, default=IQ_SUMSTATS)
    parser.add_argument("--ea4-xlsx", type=Path, default=EA4_XLSX)
    parser.add_argument("--plink-path", type=Path, default=Path("plink"))
    parser.add_argument("--polygenic-script", type=Path, default=POLYGENIC_SCRIPT)
    parser.add_argument("--gwas-p-threshold", type=float, default=1e-8)
    parser.add_argument("--clump-kb", type=int, default=200)
    parser.add_argument("--clump-r2", type=float, default=0.4)
    parser.add_argument("--n-permutations", type=int, default=100000)
    parser.add_argument("--block-size", type=int, default=250)
    parser.add_argument("--jobs", type=int, default=8)
    parser.add_argument("--seed", type=int, default=1)
    return parser.parse_args()


def validate_inputs(args: argparse.Namespace) -> None:
    for path in [
        args.sample_info_dir,
        args.aadr_ind,
        args.aadr_snp,
        args.aadr_geno,
        args.iq_sumstats,
        args.ea4_xlsx,
        args.polygenic_script,
    ]:
        if not path.exists():
            raise FileNotFoundError(f"Missing required input: {path}")


def pick_zip_member(path: Path) -> str:
    with zipfile.ZipFile(path) as zf:
        members = [
            name
            for name in zf.namelist()
            if not name.endswith("/") and "README" not in name and "CHECKSUM" not in name
        ]
        if len(members) != 1:
            raise ValueError(f"Expected one data member in {path}, found {members}")
        return members[0]


def load_iq_full_hits(path: Path, aadr_variant_ids: set[str], p_threshold: float) -> list[Hit]:
    hits: list[Hit] = []
    member = pick_zip_member(path)
    with zipfile.ZipFile(path) as zf, zf.open(member) as raw:
        handle = io.TextIOWrapper(raw, encoding="utf-8", errors="replace", newline="")
        reader = csv.DictReader(handle, delimiter="\t")
        for row in reader:
            variant_id = clean_value(row.get("SNP"))
            if not variant_id or variant_id not in aadr_variant_ids:
                continue
            try:
                p_value = float_value(row["P"])
                z_score = float_value(row["Zscore"])
            except (KeyError, TypeError, ValueError):
                continue
            if not math.isfinite(p_value) or p_value > p_threshold:
                continue
            beta_raw = row.get("stdBeta")
            effect_value = z_score if beta_raw in (None, "", "NA") else float_value(beta_raw)
            a1 = normalize_allele(row.get("A1"))
            a2 = normalize_allele(row.get("A2"))
            if not a1 or not a2:
                continue
            hits.append(
                Hit(
                    trait="iq",
                    variant_id=variant_id,
                    chrom=clean_value(row.get("CHR")),
                    position=int(round(float_value(row["POS"]))),
                    effect_allele=a1 if z_score >= 0 else a2,
                    other_allele=a2 if z_score >= 0 else a1,
                    effect_value=effect_value,
                    p_value=p_value,
                    source_gwas="Savage et al. 2018 intelligence GWAS (full summary statistics)",
                )
            )
    return hits


def workbook_rows(path: Path, sheet_name: str, header_row: int) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        worksheet = workbook[sheet_name]
    except KeyError as exc:
        raise KeyError(f"Worksheet {sheet_name!r} not found in {path}") from exc
    rows = list(worksheet.iter_rows(values_only=True))
    headers = [clean_value(value) for value in rows[header_row - 1]]
    out: list[dict[str, object]] = []
    for values in rows[header_row:]:
        out.append({headers[idx]: values[idx] if idx < len(values) else None for idx in range(len(headers))})
    return out


def load_ea4_lead_hits(path: Path, aadr_variant_ids: set[str], p_threshold: float) -> list[Hit]:
    rows = workbook_rows(path, "2. EduYears Lead SNPs", header_row=2)
    hits: list[Hit] = []
    for row in rows:
        variant_id = clean_value(row.get("SNP"))
        if not variant_id or variant_id not in aadr_variant_ids:
            continue
        beta_raw = row.get("Effect size")
        if beta_raw in (None, ""):
            continue
        beta = float_value(beta_raw)
        p_value = float_value(row["P-value"])
        if not math.isfinite(p_value) or p_value > p_threshold:
            continue
        allele1 = normalize_allele(row.get("Allele 1"))
        allele2 = normalize_allele(row.get("Allele2"))
        if not allele1 or not allele2:
            continue
        hits.append(
            Hit(
                trait="ea4",
                variant_id=variant_id,
                chrom=clean_value(row.get("Chr")),
                position=int(round(float_value(row["Position"]))),
                effect_allele=allele1 if beta >= 0 else allele2,
                other_allele=allele2 if beta >= 0 else allele1,
                effect_value=beta,
                p_value=p_value,
                source_gwas=(
                    "Lee et al. 2018 educational attainment GWAS "
                    "(supplementary lead-hit table)"
                ),
            )
        )
    return hits


def write_trait_sources(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["trait", "source_mode", "source_description"], delimiter="\t")
        writer.writeheader()
        writer.writerows(rows)


def write_results_md(
    path: Path,
    candidate_hits: list[Hit],
    clumped_hits: list[Hit],
    summary_path: Path,
    sensitivity_path: Path,
) -> None:
    candidate_counts = Counter(hit.trait for hit in candidate_hits)
    clumped_counts = Counter(hit.trait for hit in clumped_hits)
    with summary_path.open() as handle:
        summary = {(row["analysis"], row["trait"]): row for row in csv.DictReader(handle, delimiter="\t")}
    with sensitivity_path.open() as handle:
        sensitivity = {row["trait"]: row for row in csv.DictReader(handle, delimiter="\t")}
    lines = [
        "# Cognitive Comparison Results",
        "",
        "This run applies the same Colbran-style workflow used elsewhere in this repository to two cognitive comparison traits:",
        "",
        "- `iq`: Savage et al. 2018 intelligence GWAS (full summary statistics)",
        "- `ea4`: Lee et al. 2018 educational attainment GWAS lead-hit supplementary table",
        "",
        "## Lead-locus counts",
        "",
        "| Trait | Candidate 1240k hits pre-clump | Clumped lead loci |",
        "| --- | ---: | ---: |",
    ]
    for trait in TRAITS:
        lines.append(f"| `{trait}` | `{candidate_counts[trait]:,}` | `{clumped_counts[trait]:,}` |")
    lines.extend(
        [
            "",
            "## Headline sign-only results",
            "",
            f"- `iq joint_non_eur = {summary[('joint_non_eur', 'iq')]['directional_p']}`",
            f"- `ea4 joint_non_eur = {summary[('joint_non_eur', 'ea4')]['directional_p']}`",
            f"- `iq joint_non_eur_sensitivity = {sensitivity['iq']['directional_p']}`",
            f"- `ea4 joint_non_eur_sensitivity = {sensitivity['ea4']['directional_p']}`",
            "",
            "## Output",
            "",
            f"- Main summary: `{summary_path.name}`",
            "- Figure 6-style overlay can be generated with `scripts/plot_figure6_comparison.py`.",
        ]
    )
    path.write_text("\n".join(lines) + "\n")


def main() -> int:
    args = parse_args()
    validate_inputs(args)
    args.output_dir.mkdir(parents=True, exist_ok=True)

    aadr_variant_ids = load_aadr_variant_ids(args.aadr_snp)
    candidate_hits = deduplicate_hits(
        load_iq_full_hits(args.iq_sumstats, aadr_variant_ids, args.gwas_p_threshold)
        + load_ea4_lead_hits(args.ea4_xlsx, aadr_variant_ids, args.gwas_p_threshold)
    )
    if not candidate_hits:
        raise RuntimeError("No IQ or EA4 loci survived AADR intersection and genome-wide filtering.")

    write_hits(args.output_dir / "candidate_lead_variants.tsv", candidate_hits)
    write_trait_sources(
        args.output_dir / "trait_sources.tsv",
        [
            {
                "trait": "iq",
                "source_mode": "full_summary_statistics",
                "source_description": "Savage et al. 2018 intelligence GWAS (full summary statistics)",
            },
            {
                "trait": "ea4",
                "source_mode": "lead_hits",
                "source_description": "Lee et al. 2018 educational attainment GWAS lead-hit supplementary table",
            },
        ],
    )

    sample_ids, iid_to_index, canonical_to_ids = load_aadr_ind(args.aadr_ind)
    groups_by_region, union_indices = parse_region_groups(args.sample_info_dir, iid_to_index, canonical_to_ids)
    aadr_matches = load_target_variants(args.aadr_snp, {hit.variant_id for hit in candidate_hits})
    matched_preclump = build_matched_hits(candidate_hits, aadr_matches)
    if not matched_preclump:
        raise RuntimeError("No IQ or EA4 loci could be harmonized to AADR.")
    write_match_summary(args.output_dir / "candidate_match_summary.tsv", candidate_hits, matched_preclump, aadr_matches)

    eur_sample_info = args.sample_info_dir / "eur_sample_info.txt"
    eur_indices = load_eur_reference_indices(eur_sample_info, iid_to_index, canonical_to_ids)
    eur_variants = sorted({match.variant for match in matched_preclump}, key=lambda variant: variant.snp_index)
    ref_prefix, plink_ids = build_plink_reference(
        args.output_dir / "eur_reference",
        args.aadr_geno,
        len(sample_ids),
        eur_indices,
        [sample_ids[idx] for idx in eur_indices],
        eur_variants,
        args.plink_path,
    )
    selected_indices, _ = run_plink_clump(
        matched_preclump,
        ref_prefix,
        plink_ids,
        args.output_dir / "plink_clump",
        args.clump_kb,
        args.clump_r2,
        args.plink_path,
    )
    clumped_matches = [match for match in matched_preclump if match.variant.snp_index in selected_indices]
    if not clumped_matches:
        raise RuntimeError("PLINK clumping removed every IQ and EA4 locus.")
    clumped_hits = [match.hit for match in clumped_matches]
    write_hits(args.output_dir / "lead_variants.tsv", clumped_hits)
    write_match_summary(args.output_dir / "lead_variant_match_summary.tsv", clumped_hits, clumped_matches, aadr_matches)

    counts = aggregate_variant_counts(args.aadr_geno, len(sample_ids), union_indices, clumped_matches, groups_by_region)
    admixture_paths = write_admixture_files(args.output_dir, groups_by_region)
    counts_paths = write_counts_files(args.output_dir, clumped_matches, counts, groups_by_region)
    summary_path = run_sign_suite(
        args.output_dir,
        counts_paths,
        admixture_paths,
        TRAITS,
        args.polygenic_script,
        args.n_permutations,
        args.block_size,
        args.seed,
        args.jobs,
    )
    run_joint_sensitivity(
        args.output_dir,
        counts_paths,
        admixture_paths,
        TRAITS,
        args.polygenic_script,
        args.n_permutations,
        args.block_size,
        args.seed,
        args.jobs,
    )
    write_results_md(
        args.output_dir / "RESULTS.md",
        candidate_hits,
        clumped_hits,
        summary_path,
        args.output_dir / "polygenic_joint_non_eur_sensitivity.tsv",
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
